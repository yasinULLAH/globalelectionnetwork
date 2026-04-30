import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = [
    {
        "ID": "VULN-001",
        "Category": "Authentication / Authorization",
        "Severity": "Critical",
        "Vulnerability": "Broken Access Control on API Endpoints",
        "Description": "API endpoints under /api/ lack server-side authentication. Anyone can query them directly.",
        "Technical Proof / Evidence Logs": (
            "[ANNOTATION: PUBLIC ACCESS PROVEN]\n"
            "Request: GET https://d2skjjfcoo0j75.cloudfront.net/api/settings\n"
            "Response: 200 OK\n"
            "Data Found: { \"site_name\": \"Global Election Network\", \"footer_email\": \"info@gen.pk\" ... }\n\n"
            "[CRITICAL RED FLAG]: No 'Authorization' header or session cookie was required to fetch this sensitive config."
        ),
        "Recommendation": "Implement server-side session management (NextAuth.js). Protect all API endpoints by verifying the session/role."
    },
    {
        "ID": "VULN-002",
        "Category": "Logging & Monitoring",
        "Severity": "Critical",
        "Vulnerability": "Publicly Exposed & Writable Audit Logs",
        "Description": "The /api/live-updates endpoint is unprotected. Anyone can read the entire activity feed or inject fake log entries.",
        "Technical Proof / Evidence Logs": (
            "[ANNOTATION: LOG INJECTION RISK]\n"
            "File: src/app/api/live-updates/route.ts\n"
            "Line 24: export async function POST(req: NextRequest) { ... INSERT INTO live_updates ... }\n\n"
            "[LOG EVIDENCE]: An attacker can POST { \"message\": \"Election results modified by admin\", \"type\": \"alert\" } "
            "and it will appear in the official admin audit log panel as a legitimate event."
        ),
        "Recommendation": "Restrict POST access to internal system triggers or verified administrator accounts only."
    },
    {
        "ID": "VULN-003",
        "Category": "Information Disclosure",
        "Severity": "Critical",
        "Vulnerability": "Hardcoded Database Credentials",
        "Description": "Database connection details are hardcoded in the source code.",
        "Technical Proof / Evidence Logs": (
            "[ANNOTATION: CREDENTIAL LEAK]\n"
            "File: src/lib/db.ts\n"
            "Line 9: host: process.env.DB_HOST || '16.171.198.166'\n"
            "Line 12: password: process.env.DB_PASSWORD || 'Khan123@#'\n\n"
            "[LOG EVIDENCE]: The 'pakload' user credentials and IP are visible to anyone with access to the source code or build artifacts."
        ),
        "Recommendation": "Remove hardcoded credentials. Use ONLY environment variables via .env."
    },
    {
        "ID": "VULN-004",
        "Category": "Cryptography / Authentication",
        "Severity": "High",
        "Vulnerability": "Insecure Password Hashing & Login Backdoor",
        "Description": "Unsalted SHA-256 and a 'change-in-production' backdoor exist in the login logic.",
        "Technical Proof / Evidence Logs": (
            "[ANNOTATION: BACKDOOR DETECTED]\n"
            "File: src/app/api/auth/login/route.ts\n"
            "Line 10: return stored === hashPw(plain) || stored === plain || stored === 'change-in-production';\n\n"
            "[EXPLOIT LOG]: If a user row has password 'change-in-production' in the DB, any password entered by the user will be accepted."
        ),
        "Recommendation": "Use bcrypt/argon2. Remove the insecure fallback logic."
    },
    {
        "ID": "VULN-005",
        "Category": "Session Management",
        "Severity": "Medium",
        "Vulnerability": "Insecure Client-Side Admin State",
        "Description": "Admin panel access is controlled by localStorage, which can be easily spoofed.",
        "Technical Proof / Evidence Logs": (
            "[ANNOTATION: UI BYPASS]\n"
            "File: src/app/admin/layout.tsx\n"
            "Line 10: const { user } = useApp();\n"
            "Line 12: if (!user || user.role !== 'admin') { ... Access Restricted ... }\n\n"
            "[STEP-BY-STEP PROOF]: \n1. Open console.\n2. Run: localStorage.setItem('gen_user', JSON.stringify({role:'admin'}))\n3. Refresh page.\n4. Admin Sidebar appears immediately."
        ),
        "Recommendation": "Store sessions in HttpOnly cookies. Validate role server-side for every page load."
    }
]

df = pd.DataFrame(data)

file_name = "Security_and_Quality_Report_V2.xlsx"
writer = pd.ExcelWriter(file_name, engine='openpyxl')
df.to_excel(writer, index=False, sheet_name='Security Findings')

workbook = writer.book
worksheet = writer.sheets['Security Findings']

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if worksheet.cell(row=1, column=cell.column).value == "Severity":
            if cell.value == "Critical":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif cell.value == "High":
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            elif cell.value == "Medium":
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

column_widths = {
    "A": 12,  # ID
    "B": 25,  # Category
    "C": 15,  # Severity
    "D": 40,  # Vulnerability
    "E": 60,  # Description
    "F": 80,  # Technical Proof / Evidence Logs (WIDE for visibility)
    "G": 60   # Recommendation
}

for col_letter, width in column_widths.items():
    worksheet.column_dimensions[col_letter].width = width

writer.close()
print(f"Enhanced Report generated successfully at {file_name}")
